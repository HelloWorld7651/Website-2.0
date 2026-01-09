from pathlib import Path
import os, sys, argparse, shutil, re, math
import numpy as np, cv2, pdfplumber, pytesseract, pandas as pd
from PIL import Image
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment
from openpyxl.utils import get_column_letter

WORKSHEET_HEADER_FIELDS = ["Ref #","Hits","Search Query","DBs","Default Operator","Plurals","British Equivalents","Time Stamp"]
EXPECTED_COLUMN_COUNT = len(WORKSHEET_HEADER_FIELDS)
REFERENCE_NUMBER_REGEX = re.compile(r"^\s*L\d+\s*$")
INVALID_SHEET_CHARS_REGEX = re.compile(r"[:\\/?*\[\]]")


def set_tesseract_cmd():
    tesseract_executable_path = shutil.which("tesseract")
    for candidate_path in ("/opt/homebrew/bin/tesseract","/usr/local/bin/tesseract","/usr/bin/tesseract"):
        if not tesseract_executable_path and os.path.exists(candidate_path):
            tesseract_executable_path = candidate_path
    if not tesseract_executable_path:
        raise RuntimeError("Install Tesseract (e.g., `brew install tesseract`).")
    pytesseract.pytesseract.tesseract_cmd = tesseract_executable_path
    print("Using tesseract:", tesseract_executable_path)

def pil_image_to_cv2_bgr(pil_image):
    return cv2.cvtColor(np.array(pil_image), cv2.COLOR_RGB2BGR)


def detect_table_gridline_centers(grayscale_image, min_line_fraction):
    binary_inverted_image = cv2.adaptiveThreshold(grayscale_image,255,cv2.ADAPTIVE_THRESH_MEAN_C,cv2.THRESH_BINARY_INV,15,10)
    image_height, image_width = binary_inverted_image.shape
    horizontal_kernel = cv2.getStructuringElement(cv2.MORPH_RECT,(max(10,image_width//40),1))
    vertical_kernel = cv2.getStructuringElement(cv2.MORPH_RECT,(1,max(10,image_height//35)))
    horizontal_lines_image = cv2.dilate(cv2.erode(binary_inverted_image,horizontal_kernel,1),horizontal_kernel,1)
    vertical_lines_image = cv2.dilate(cv2.erode(binary_inverted_image,vertical_kernel,1),vertical_kernel,1)
    vertical_line_pixel_counts_per_column = vertical_lines_image.sum(axis=0)//255
    horizontal_line_pixel_counts_per_row = horizontal_lines_image.sum(axis=1)//255
    vertical_line_x_positions = np.where(vertical_line_pixel_counts_per_column > int(image_height*min_line_fraction))[0]
    horizontal_line_y_positions = np.where(horizontal_line_pixel_counts_per_row > int(image_width*min_line_fraction))[0]

    def compute_group_centers(sorted_indices):
        if sorted_indices.size == 0:
            return []
        groups = []
        start_index = previous_index = sorted_indices[0]
        for index in sorted_indices[1:]:
            if index == previous_index + 1:
                previous_index = index
            else:
                groups.append((start_index, previous_index))
                start_index = previous_index = index
        groups.append((start_index, previous_index))
        return [(a+b)//2 for a,b in groups]

    x_centers = compute_group_centers(vertical_line_x_positions)
    y_centers = compute_group_centers(horizontal_line_y_positions)
    return sorted(set(x_centers)), sorted(set(y_centers))

def merge_hyphenated_linebreaks(text):
    if not text:
        return text
    text = re.sub(r'(\w)-\n(\w)', r'\1\2', text)
    text = re.sub(r'[ \t]+\n', '\n', text)
    return text.strip()


def ocr_text_from_bgr_cell(cell_bgr_image, page_seg_mode=6):
    if cell_bgr_image.size == 0:
        return ""
    cell_gray = cv2.cvtColor(cell_bgr_image, cv2.COLOR_BGR2GRAY)
    cell_gray = cv2.fastNlMeansDenoising(cell_gray, h=15)
    _, thresholded_image = cv2.threshold(cell_gray,0,255,cv2.THRESH_BINARY+cv2.THRESH_OTSU)
    best_text = ""
    for candidate_psm in (page_seg_mode, 4, 7):
        recognized_text = pytesseract.image_to_string(
            thresholded_image, config=f"--psm {candidate_psm} -c preserve_interword_spaces=1", lang="eng"
        ).strip()
        if len(recognized_text.replace(" ","")) > len(best_text.replace(" ","")):
            best_text = recognized_text
    return merge_hyphenated_linebreaks(best_text)

def ocr_text_from_cell_with_retry(full_bgr_image, x_left, x_right, y_top, y_bottom, padding_pixels, page_seg_mode):
    image_height, image_width = full_bgr_image.shape[:2]
    for extra_padding in (0, 6, 12):
        x1 = max(0, x_left - padding_pixels - extra_padding)
        x2 = min(image_width, x_right + padding_pixels + extra_padding)
        y1 = max(0, y_top - padding_pixels - extra_padding)
        y2 = min(image_height, y_bottom + padding_pixels + extra_padding)
        text_candidate = ocr_text_from_bgr_cell(full_bgr_image[y1:y2, x1:x2], page_seg_mode=page_seg_mode)
        if text_candidate:
            return text_candidate
    return ""


def process_pdf(input_pdf_path: Path, render_dpi=450, min_line_fraction=0.38, padding_pixels=6, page_seg_mode=6):
    page_number_and_dataframe_list = []
    with pdfplumber.open(input_pdf_path) as pdf_document:
        print(f"Processing {input_pdf_path.name} ({len(pdf_document.pages)} pages)…")
        for page_number, pdf_page in enumerate(pdf_document.pages, 1):
            page_bgr_image = pil_image_to_cv2_bgr(pdf_page.to_image(resolution=render_dpi).original)
            page_grayscale_image = cv2.cvtColor(page_bgr_image, cv2.COLOR_BGR2GRAY)
            x_boundaries, y_boundaries = detect_table_gridline_centers(page_grayscale_image, min_line_fraction=min_line_fraction)
            if len(x_boundaries) < 2 or len(y_boundaries) < 2:
                full_page_text = ocr_text_from_bgr_cell(page_bgr_image, page_seg_mode=page_seg_mode)
                fallback_dataframe = pd.DataFrame([[full_page_text] + [""]*(EXPECTED_COLUMN_COUNT-1)])
                fallback_dataframe.columns = WORKSHEET_HEADER_FIELDS
                page_number_and_dataframe_list.append((page_number, fallback_dataframe))
                continue
            extracted_table_rows = []
            for row_index in range(len(y_boundaries)-1):
                y_top = y_boundaries[row_index] + 1
                y_bottom = y_boundaries[row_index+1] - 1
                row_cells_text = []
                for column_index in range(len(x_boundaries)-1):
                    x_left = x_boundaries[column_index] + 1
                    x_right = x_boundaries[column_index+1] - 1
                    cell_text = ocr_text_from_cell_with_retry(page_bgr_image, x_left, x_right, y_top, y_bottom, padding_pixels=padding_pixels, page_seg_mode=page_seg_mode)
                    row_cells_text.append(cell_text)
                extracted_table_rows.append(row_cells_text)
            page_number_and_dataframe_list.append((page_number, pd.DataFrame(extracted_table_rows)))
    return page_number_and_dataframe_list

def normalize_dataframe_to_expected_columns(dataframe):
    current_column_count = dataframe.shape[1]
    if current_column_count < EXPECTED_COLUMN_COUNT:
        dataframe = dataframe.reindex(columns=range(EXPECTED_COLUMN_COUNT), fill_value="")
    elif current_column_count > EXPECTED_COLUMN_COUNT:
        dataframe = dataframe.iloc[:, :EXPECTED_COLUMN_COUNT]
    return dataframe

def fold_continuation_rows_into_previous_reference(dataframe):
    dataframe = normalize_dataframe_to_expected_columns(dataframe)
    folded_rows = []
    for row_index in range(len(dataframe)):
        row_cells = [str(dataframe.iat[row_index, col_index] or "").strip() for col_index in range(dataframe.shape[1])]
        if REFERENCE_NUMBER_REGEX.match(row_cells[0]):
            folded_rows.append(row_cells)
        else:
            if not folded_rows:
                folded_rows.append(row_cells)
            else:
                previous_row = folded_rows[-1]
                for col_index in range(len(previous_row)):
                    if row_cells[col_index]:
                        previous_row[col_index] = (previous_row[col_index] + ("\n" if previous_row[col_index] else "") + row_cells[col_index]).strip()
                folded_rows[-1] = previous_row
    output_dataframe = pd.DataFrame(folded_rows)
    output_dataframe = normalize_dataframe_to_expected_columns(output_dataframe)
    output_dataframe.columns = WORKSHEET_HEADER_FIELDS
    return output_dataframe

def generate_unique_sheet_name(base_title, used_titles):
    sanitized_name = INVALID_SHEET_CHARS_REGEX.sub("_", base_title)[:31] or "Sheet"
    if sanitized_name not in used_titles:
        used_titles.add(sanitized_name)
        return sanitized_name
    for suffix_index in range(1, 1000):
        candidate_name = (sanitized_name[:31-len(f"_{suffix_index}")] + f"_{suffix_index}")
        if candidate_name not in used_titles:
            used_titles.add(candidate_name)
            return candidate_name
    raise RuntimeError("Sheet naming overflow.")

def autosize_columns_and_set_row_heights(worksheet, max_column_count):
    for column_number in range(1, max_column_count+1):
        max_char_len = 12
        for row_number in range(1, worksheet.max_row+1):
            cell_value = worksheet.cell(row_number, column_number).value
            if cell_value:
                max_char_len = max(max_char_len, min(60, len(str(cell_value))))
        worksheet.column_dimensions[get_column_letter(column_number)].width = max_char_len
    for row_number in range(2, worksheet.max_row+1):
        line_count = 1
        for column_number in range(1, max_column_count+1):
            cell_value = worksheet.cell(row_number, column_number).value
            if cell_value:
                line_count = max(line_count, str(cell_value).count("\n")+1)
        worksheet.row_dimensions[row_number].height = min(409, 13*line_count)

def write_workbook_with_many_sheets(pdf_to_page_dataframe_pairs, output_workbook_path: Path):
    workbook = Workbook()
    if workbook.active:
        workbook.remove(workbook.active)
    thin_side = Side(style="thin", color="000000")
    cell_border = Border(left=thin_side,right=thin_side,top=thin_side,bottom=thin_side)
    wrap_alignment = Alignment(wrap_text=True, vertical="top")
    used_sheet_names = set()

    for pdf_file_path, page_dataframe_list in pdf_to_page_dataframe_pairs:
        normalized_page_tables = []
        for _, page_dataframe in page_dataframe_list:
            if page_dataframe.empty:
                continue
            if list(page_dataframe.columns) == WORKSHEET_HEADER_FIELDS:
                normalized_page_tables.append(page_dataframe)
            else:
                normalized_page_tables.append(fold_continuation_rows_into_previous_reference(page_dataframe))
        if not normalized_page_tables:
            normalized_page_tables = [pd.DataFrame(columns=WORKSHEET_HEADER_FIELDS)]
        combined_table = pd.concat(normalized_page_tables, ignore_index=True)

        worksheet = workbook.create_sheet(title=generate_unique_sheet_name(pdf_file_path.stem, used_sheet_names))
        for column_index, header_title in enumerate(WORKSHEET_HEADER_FIELDS, 1):
            header_cell = worksheet.cell(1, column_index, header_title)
            header_cell.alignment = wrap_alignment
            header_cell.border = cell_border
        worksheet.freeze_panes = "A2"
        for row_index in range(combined_table.shape[0]):
            for column_index in range(combined_table.shape[1]):
                data_cell = worksheet.cell(row_index+2, column_index+1, combined_table.iat[row_index, column_index])
                data_cell.alignment = wrap_alignment
                data_cell.border = cell_border
        autosize_columns_and_set_row_heights(worksheet, EXPECTED_COLUMN_COUNT)

    workbook.save(output_workbook_path)
    return output_workbook_path

def main():
    argument_parser = argparse.ArgumentParser()
    argument_parser.add_argument("--inputs", nargs="*", type=Path, default=[])
    argument_parser.add_argument("--dir", type=Path)
    argument_parser.add_argument("--glob", type=str, default="*.pdf")
    argument_parser.add_argument("--out", type=Path, required=True)
    argument_parser.add_argument("--dpi", type=int, default=450)
    argument_parser.add_argument("--min-line-frac", type=float, default=0.38)
    argument_parser.add_argument("--pad", type=int, default=6)
    argument_parser.add_argument("--psm", type=int, default=6)
    cli_args = argument_parser.parse_args()

    set_tesseract_cmd()

    discovered_pdf_paths = []
    if cli_args.dir:
        discovered_pdf_paths += sorted(cli_args.dir.glob(cli_args.glob))
    if cli_args.inputs:
        discovered_pdf_paths += cli_args.inputs
    seen_paths = set()
    ordered_unique_paths = []
    for pdf_file_path in discovered_pdf_paths:
        absolute_pdf_path = pdf_file_path.resolve()
        if absolute_pdf_path not in seen_paths:
            seen_paths.add(absolute_pdf_path)
            ordered_unique_paths.append(absolute_pdf_path)
    if not ordered_unique_paths:
        sys.exit("No PDFs found.")
    for absolute_pdf_path in ordered_unique_paths:
        if not absolute_pdf_path.exists():
            sys.exit(f"Missing: {absolute_pdf_path}")

    pdf_to_page_dataframe_pairs = []
    for absolute_pdf_path in ordered_unique_paths:
        page_dataframes = process_pdf(
            absolute_pdf_path,
            render_dpi=cli_args.dpi,
            min_line_fraction=cli_args.min_line_frac,
            padding_pixels=cli_args.pad,
            page_seg_mode=cli_args.psm
        )
        pdf_to_page_dataframe_pairs.append((absolute_pdf_path, page_dataframes))
        print(" ->", absolute_pdf_path.name, "done")

    saved_workbook_path = write_workbook_with_many_sheets(pdf_to_page_dataframe_pairs, cli_args.out)
    print("Saved workbook:", saved_workbook_path)

if __name__ == "__main__":
    main()
