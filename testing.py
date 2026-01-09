#helped by the internet, gpt, daniel's spite

'''
input: given folder, search for .pdf, output destination, factors
dpi: higher better detection slower speed
minline: detect line
pad: crop includes boarder stuff

collect pdf
red page as image
find table
tesseract ocr each table grid
write excel
'''
from pathlib import Path
import os, sys, argparse, shutil, re, math
import numpy as np, cv2, pdfplumber, pytesseract, pandas as pd
from PIL import Image
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment
from openpyxl.utils import get_column_letter
HEADER = ["Ref #","Hits","Search Query","DBs","Default Operator","Plurals","British Equivalents","Time Stamp"]
EXPECTED_COLS = len(HEADER)
REF_RE = re.compile(r"^\s*L\d+\s*$")
INVALID_SHEET = re.compile(r"[:\\/?*\[\]]")


def set_tesseract_cmd():
    tes = shutil.which("tesseract")
    for p in ("/opt/homebrew/bin/tesseract","/usr/local/bin/tesseract","/usr/bin/tesseract"):
        if not tes and os.path.exists(p): tes = p
    if not tes: raise RuntimeError("Install Tesseract (e.g., `brew install tesseract`).")
    pytesseract.pytesseract.tesseract_cmd = tes
    print("Using tesseract:", tes)

def pil_to_cv(img): return cv2.cvtColor(np.array(img), cv2.COLOR_RGB2BGR)


def detect_lines(gray, min_line_frac):
    bw = cv2.adaptiveThreshold(gray,255,cv2.ADAPTIVE_THRESH_MEAN_C,cv2.THRESH_BINARY_INV,15,10)
    h,w = bw.shape
    hk = cv2.getStructuringElement(cv2.MORPH_RECT,(max(10,w//40),1))
    vk = cv2.getStructuringElement(cv2.MORPH_RECT,(1,max(10,h//35)))
    horiz = cv2.dilate(cv2.erode(bw,hk,1),hk,1)
    vert  = cv2.dilate(cv2.erode(bw,vk,1),vk,1)
    col_sum = vert.sum(axis=0)//255; row_sum = horiz.sum(axis=1)//255
    vx = np.where(col_sum > int(h*min_line_frac))[0]
    hy = np.where(row_sum > int(w*min_line_frac))[0]

    def centers(idx):
        if idx.size==0: return []
        groups=[]; s=prev=idx[0]
        for i in idx[1:]:
            if i==prev+1: prev=i
            else: groups.append((s,prev)); s=prev=i
        groups.append((s,prev))
        return [(a+b)//2 for a,b in groups]
    return sorted(set(centers(vx))), sorted(set(centers(hy)))

def _fix_hyphens(s:str)->str:
    if not s: return s
    s = re.sub(r'(\w)-\n(\w)', r'\1\2', s)  
    s = re.sub(r'[ \t]+\n', '\n', s)         
    return s.strip()



def ocr_cell(img_bgr, psm=6):
    if img_bgr.size==0: return ""
    g = cv2.cvtColor(img_bgr, cv2.COLOR_BGR2GRAY)
    g = cv2.fastNlMeansDenoising(g, h=15)
    _,th = cv2.threshold(g,0,255,cv2.THRESH_BINARY+cv2.THRESH_OTSU)
    best = ""
    for p in (psm, 4, 7):  
        txt = pytesseract.image_to_string(
            th, config=f"--psm {p} -c preserve_interword_spaces=1", lang="eng"
        ).strip()
        if len(txt.replace(" ","")) > len(best.replace(" ","")):
            best = txt
    return _fix_hyphens(best)

def ocr_cell_with_retry(cv_img, x1,x2,y1,y2, pad, psm):
    H,W = cv_img.shape[:2]
    for extra in (0, 6, 12):
        xa = max(0, x1 - pad - extra); xb = min(W, x2 + pad + extra)
        ya = max(0, y1 - pad - extra); yb = min(H, y2 + pad + extra)
        txt = ocr_cell(cv_img[ya:yb, xa:xb], psm=psm)
        if txt: return txt
    return ""


def process_pdf(pdf_path: Path, dpi=450, min_line_frac=0.38, pad=6, psm=6):
    pages=[]
    with pdfplumber.open(pdf_path) as doc:
        print(f"Processing {pdf_path.name} ({len(doc.pages)} pages)…")
        for n, p in enumerate(doc.pages, 1):
            cv = pil_to_cv(p.to_image(resolution=dpi).original)
            gray = cv2.cvtColor(cv, cv2.COLOR_BGR2GRAY)
            xs, ys = detect_lines(gray, min_line_frac=min_line_frac)
            if len(xs)<2 or len(ys)<2:
                text = ocr_cell(cv, psm=psm)
                tmp = pd.DataFrame([[text] + [""]*(EXPECTED_COLS-1)])
                tmp.columns = HEADER
                pages.append((n, tmp)); continue
            rows=[]
            for r in range(len(ys)-1):
                y1,y2 = ys[r]+1, ys[r+1]-1
                row=[]
                for c in range(len(xs)-1):
                    x1,x2 = xs[c]+1, xs[c+1]-1
                    row.append(ocr_cell_with_retry(cv, x1,x2,y1,y2, pad=pad, psm=psm))
                rows.append(row)
            pages.append((n, pd.DataFrame(rows)))
    return pages

def normalize_columns(df):
    cols = df.shape[1]
    if cols < EXPECTED_COLS:
        df = df.reindex(columns=range(EXPECTED_COLS), fill_value="")
    elif cols > EXPECTED_COLS:
        df = df.iloc[:, :EXPECTED_COLS]
    return df

def fold_continuations(df):
    df = normalize_columns(df)
    out=[]
    for r in range(len(df)):
        cells = [str(df.iat[r,c] or "").strip() for c in range(df.shape[1])]
        if REF_RE.match(cells[0]):
            out.append(cells)
        else:
            if not out:
                out.append(cells)
            else:
                prev = out[-1]
                for c in range(len(prev)):
                    if cells[c]:
                        prev[c] = (prev[c] + ("\n" if prev[c] else "") + cells[c]).strip()
                out[-1] = prev
    out_df = pd.DataFrame(out)
    out_df = normalize_columns(out_df)
    out_df.columns = HEADER
    return out_df

def unique_sheet_name(base, used):
    name = INVALID_SHEET.sub("_", base)[:31] or "Sheet"
    if name not in used:
        used.add(name); return name
    for i in range(1, 1000):
        cand = (name[:31-len(f"_{i}")] + f"_{i}")
        if cand not in used:
            used.add(cand); return cand
    raise RuntimeError("Sheet naming overflow.")

def autosize_and_rowheights(ws, max_cols):
    for col in range(1, max_cols+1):
        maxlen = 12
        for row in range(1, ws.max_row+1):
            v = ws.cell(row,col).value
            if v: maxlen = max(maxlen, min(60, len(str(v))))
        ws.column_dimensions[get_column_letter(col)].width = maxlen
    for row in range(2, ws.max_row+1):
        lines = 1
        for col in range(1, max_cols+1):
            v = ws.cell(row,col).value
            if v:
                lines = max(lines, str(v).count("\n")+1)
        ws.row_dimensions[row].height = min(409, 13*lines) 

def write_many_sheets(pdf_to_pages, out_xlsx: Path):
    wb = Workbook()
    if wb.active: wb.remove(wb.active)
    thin = Side(style="thin", color="000000")
    border = Border(left=thin,right=thin,top=thin,bottom=thin)
    wrap = Alignment(wrap_text=True, vertical="top")
    used = set()

    for pdf_path, pages in pdf_to_pages:
        logical=[]
        for _, df in pages:
            if df.empty: continue
            if list(df.columns)==HEADER:
                logical.append(df)
            else:
                logical.append(fold_continuations(df))
        if not logical:
            logical=[pd.DataFrame(columns=HEADER)]
        combined = pd.concat(logical, ignore_index=True)

        ws = wb.create_sheet(title=unique_sheet_name(pdf_path.stem, used))
        for j,h in enumerate(HEADER,1):
            cell = ws.cell(1,j,h); cell.alignment=wrap
            cell.border=border
        ws.freeze_panes="A2"
        for i in range(combined.shape[0]):
            for j in range(combined.shape[1]):
                cell = ws.cell(i+2,j+1,combined.iat[i,j])
                cell.alignment=wrap; cell.border=border
        autosize_and_rowheights(ws, EXPECTED_COLS)

    wb.save(out_xlsx); return out_xlsx

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--inputs", nargs="*", type=Path, default=[])
    ap.add_argument("--dir", type=Path)
    ap.add_argument("--glob", type=str, default="*.pdf")
    ap.add_argument("--out", type=Path, required=True)
    ap.add_argument("--dpi", type=int, default=450)
    ap.add_argument("--min-line-frac", type=float, default=0.38)
    ap.add_argument("--pad", type=int, default=6)
    ap.add_argument("--psm", type=int, default=6)
    args = ap.parse_args()

    set_tesseract_cmd()

    pdfs=[]
    if args.dir: pdfs += sorted(args.dir.glob(args.glob))
    if args.inputs: pdfs += args.inputs
    seen=set(); ordered=[]
    for p in pdfs:
        p=p.resolve()
        if p not in seen:
            seen.add(p); ordered.append(p)
    if not ordered: sys.exit("No PDFs found.")
    for p in ordered:
        if not p.exists(): sys.exit(f"Missing: {p}")

    bundle=[]
    for p in ordered:
        pages = process_pdf(p, dpi=args.dpi, min_line_frac=args.min_line_frac, pad=args.pad, psm=args.psm)
        bundle.append((p,pages))
        print(" ->", p.name, "done")

    out = write_many_sheets(bundle, args.out)
    print("Saved workbook:", out)

if __name__ == "__main__":
    main()


'''
python3 testing.py \
  --dir "/Users/helloworld/Downloads/pe2e" \
  --glob "*.pdf" \
  --out "/Users/helloworld/Downloads/pe2e_all.xlsx" \
  --dpi 500 --min-line-frac 0.35 --pad 10 --psm 6
'''