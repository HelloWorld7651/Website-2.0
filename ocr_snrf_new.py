
"""
input: given folder, search for .pdf, output destination, factors
dpi: higher dpi -> better detection, slower speed
min-line-frac: fraction of page that must be "line" pixels to count as a grid line
pad: extra padding around each detected cell for OCR

Pipeline:
- collect PDFs
- render each page as image
- find table grid
- OCR each table cell with Tesseract
- fold multi-line logical rows
- write one Excel workbook with one sheet per PDF
"""

from pathlib import Path
import os
import sys
import argparse
import shutil
import re
import math

import numpy as np
import cv2
import pdfplumber
import pytesseract
import pandas as pd
from PIL import Image

from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

HEADER = [
    "Ref #",
    "Hits",
    "Search Query",
    "DBs",
    "Default Operator",
    "Plurals",
    "British Equivalents",
    "Time Stamp",
]
EXPECTED_COLS = len(HEADER)

# More forgiving: L1, L 1, L1/, L1- all count as "ref rows"
REF_RE = re.compile(r"^\s*L\s*\d+\s*[/\-]?\s*$")

INVALID_SHEET = re.compile(r"[:\\/?*\[\]]")


# ---------------------------------------------------------------------------
# Tesseract setup
# ---------------------------------------------------------------------------

def set_tesseract_cmd():
    tes = shutil.which("tesseract")
    for p in ("/opt/homebrew/bin/tesseract",
              "/usr/local/bin/tesseract",
              "/usr/bin/tesseract"):
        if not tes and os.path.exists(p):
            tes = p
    if not tes:
        raise RuntimeError("Install Tesseract (e.g., `brew install tesseract`).")
    pytesseract.pytesseract.tesseract_cmd = tes
    print("Using tesseract:", tes)


# ---------------------------------------------------------------------------
# Image helpers
# ---------------------------------------------------------------------------

def pil_to_cv(img: Image.Image) -> np.ndarray:
    """Convert a PIL image to OpenCV BGR array."""
    return cv2.cvtColor(np.array(img), cv2.COLOR_RGB2BGR)


def detect_lines(gray: np.ndarray, min_line_frac: float):
    """
    Detect vertical and horizontal table lines using morphology.
    Returns:
        xs: sorted list of x positions of vertical lines
        ys: sorted list of y positions of horizontal lines
    """
    # Binary inverse: table lines become white on black
    bw = cv2.adaptiveThreshold(
        gray,
        255,
        cv2.ADAPTIVE_THRESH_MEAN_C,
        cv2.THRESH_BINARY_INV,
        15,
        10,
    )

    h, w = bw.shape
    # Horizontal & vertical kernels
    hk = cv2.getStructuringElement(cv2.MORPH_RECT, (max(10, w // 40), 1))
    vk = cv2.getStructuringElement(cv2.MORPH_RECT, (1, max(10, h // 35)))

    horiz = cv2.dilate(cv2.erode(bw, hk, 1), hk, 1)
    vert = cv2.dilate(cv2.erode(bw, vk, 1), vk, 1)

    col_sum = vert.sum(axis=0) // 255
    row_sum = horiz.sum(axis=1) // 255

    vx = np.where(col_sum > int(h * min_line_frac))[0]
    hy = np.where(row_sum > int(w * min_line_frac))[0]

    def centers(idx):
        if idx.size == 0:
            return []
        groups = []
        s = prev = idx[0]
        for i in idx[1:]:
            if i == prev + 1:
                prev = i
            else:
                groups.append((s, prev))
                s = prev = i
        groups.append((s, prev))
        return [(a + b) // 2 for a, b in groups]

    return sorted(set(centers(vx))), sorted(set(centers(hy)))


# ---------------------------------------------------------------------------
# OCR helpers
# ---------------------------------------------------------------------------

def _fix_hyphens(s: str) -> str:
    """Join hyphenated words that break across lines, trim whitespace."""
    if not s:
        return s
    # e.g., "infor-\nmation" -> "information"
    s = re.sub(r"(\w)-\n(\w)", r"\1\2", s)
    # trim spaces before newline
    s = re.sub(r"[ \t]+\n", "\n", s)
    return s.strip()


def ocr_cell(img_bgr: np.ndarray, psm: int = 6) -> str:
    """OCR a cropped cell; try a couple of PSMs and pick the 'densest' result."""
    if img_bgr.size == 0:
        return ""

    g = cv2.cvtColor(img_bgr, cv2.COLOR_BGR2GRAY)
    g = cv2.fastNlMeansDenoising(g, h=15)
    _, th = cv2.threshold(g, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)

    best = ""
    for p in (psm, 4, 7):
        txt = pytesseract.image_to_string(
            th,
            config=f"--psm {p} -c preserve_interword_spaces=1",
            lang="eng",
        ).strip()
        # keep the version with most non-space characters
        if len(txt.replace(" ", "")) > len(best.replace(" ", "")):
            best = txt

    return _fix_hyphens(best)


def ocr_cell_with_retry(
    cv_img: np.ndarray,
    x1: int,
    x2: int,
    y1: int,
    y2: int,
    pad: int,
    psm: int,
) -> str:
    """
    Try OCR with increasing padding around the cell in case the grid is slightly off.
    """
    H, W = cv_img.shape[:2]
    for extra in (0, 6, 12):
        xa = max(0, x1 - pad - extra)
        xb = min(W, x2 + pad + extra)
        ya = max(0, y1 - pad - extra)
        yb = min(H, y2 + pad + extra)
        txt = ocr_cell(cv_img[ya:yb, xa:xb], psm=psm)
        if txt:
            return txt
    return ""


# ---------------------------------------------------------------------------
# Per-PDF processing
# ---------------------------------------------------------------------------

def process_pdf(
    pdf_path: Path,
    dpi: int = 450,
    min_line_frac: float = 0.38,
    pad: int = 6,
    psm: int = 6,
):
    """
    Process a single PDF into a list of (page_number, DataFrame) pairs.
    DataFrames may be raw OCR grids or already-structured tables.
    """
    pages = []
    with pdfplumber.open(pdf_path) as doc:
        print(f"Processing {pdf_path.name} ({len(doc.pages)} pages)…")
        for n, p in enumerate(doc.pages, 1):
            cv_img = pil_to_cv(p.to_image(resolution=dpi).original)
            gray = cv2.cvtColor(cv_img, cv2.COLOR_BGR2GRAY)
            xs, ys = detect_lines(gray, min_line_frac=min_line_frac)

            # ------------------------------------------------------------------
            # CASE 1: No reliable grid detected -> treat page text as lines and
            # split logical rows whenever we see "L<number>" at the start of a line.
            # ------------------------------------------------------------------
            if len(xs) < 2 or len(ys) < 2:
                full_text = ocr_cell(cv_img, psm=psm)
                lines = [ln.strip() for ln in full_text.splitlines() if ln.strip()]

                rows = []
                # current logical row
                cur = [""] * EXPECTED_COLS  # [Ref#, Hits, Search Query, DBs, ...]

                for ln in lines:
                    m = re.match(r"^L\s*\d+\b", ln)
                    if m:
                        # start a new logical row
                        if any(cur):
                            rows.append(cur)
                        ref = m.group(0).replace(" ", "")  # "L 1" -> "L1"
                        rest = ln[m.end():].strip()
                        cur = [ref, "", rest, "", "", "", "", ""]
                    else:
                        # continuation -> append to Search Query
                        if not any(cur):
                            # first row on page had no "L<number>"
                            cur = ["", "", ln, "", "", "", "", ""]
                        else:
                            cur[2] = (
                                cur[2]
                                + ("\n" if cur[2] else "")
                                + ln
                            ).strip()

                if any(cur):
                    rows.append(cur)
                if not rows:
                    # fall back: one giant cell in first column
                    rows = [[full_text] + [""] * (EXPECTED_COLS - 1)]

                tmp = pd.DataFrame(rows, columns=HEADER)
                pages.append((n, tmp))
                continue

            # ------------------------------------------------------------------
            # CASE 2: Normal grid detected -> OCR each cell in the grid
            # ------------------------------------------------------------------
            rows = []
            for r in range(len(ys) - 1):
                y1, y2 = ys[r] + 1, ys[r + 1] - 1
                row = []
                for c in range(len(xs) - 1):
                    x1, x2 = xs[c] + 1, xs[c + 1] - 1
                    row.append(
                        ocr_cell_with_retry(
                            cv_img,
                            x1,
                            x2,
                            y1,
                            y2,
                            pad=pad,
                            psm=psm,
                        )
                    )
                rows.append(row)
            pages.append((n, pd.DataFrame(rows)))

    return pages


# ---------------------------------------------------------------------------
# Table cleanup
# ---------------------------------------------------------------------------

def normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Ensure DataFrame has exactly EXPECTED_COLS columns."""
    cols = df.shape[1]
    if cols < EXPECTED_COLS:
        df = df.reindex(columns=range(EXPECTED_COLS), fill_value="")
    elif cols > EXPECTED_COLS:
        df = df.iloc[:, :EXPECTED_COLS]
    return df


def fold_continuations(df: pd.DataFrame) -> pd.DataFrame:
    """
    Fold continuation rows into the previous logical row.

    Rules:
    - If first cell matches REF_RE (L1, L 2, L3/, etc) => new logical row.
    - If first cell is BLANK and there is a previous row => continuation.
    - Otherwise keep as its own row (don't merge weird junk over an L-row).
    """
    df = normalize_columns(df)
    out = []

    for r in range(len(df)):
        cells = [str(df.iat[r, c] or "").strip() for c in range(df.shape[1])]
        ref = cells[0]

        if REF_RE.match(ref):
            # clearly a new L<number> row
            out.append(cells)
        elif not ref.strip() and out:
            # empty ref -> continuation of previous row
            prev = out[-1]
            for c in range(len(prev)):
                if cells[c]:
                    prev[c] = (
                        prev[c]
                        + ("\n" if prev[c] else "")
                        + cells[c]
                    ).strip()
            out[-1] = prev
        else:
            # weird but non-empty first column: keep as its own row
            out.append(cells)

    out_df = pd.DataFrame(out)
    out_df = normalize_columns(out_df)
    out_df.columns = HEADER
    return out_df


# ---------------------------------------------------------------------------
# Excel helpers
# ---------------------------------------------------------------------------

def unique_sheet_name(base: str, used: set) -> str:
    """Generate a unique, Excel-safe sheet name from base name."""
    name = INVALID_SHEET.sub("_", base)[:31] or "Sheet"
    if name not in used:
        used.add(name)
        return name

    for i in range(1, 1000):
        cand = name[: 31 - len(f"_{i}")] + f"_{i}"
        if cand not in used:
            used.add(cand)
            return cand

    raise RuntimeError("Sheet naming overflow.")


def autosize_and_rowheights(ws, max_cols: int):
    """Auto size columns and row heights based on content length and line breaks."""
    # columns
    for col in range(1, max_cols + 1):
        maxlen = 12
        for row in range(1, ws.max_row + 1):
            v = ws.cell(row, col).value
            if v:
                maxlen = max(maxlen, min(60, len(str(v))))
        ws.column_dimensions[get_column_letter(col)].width = maxlen

    # rows (start from 2 to skip header)
    for row in range(2, ws.max_row + 1):
        lines = 1
        for col in range(1, max_cols + 1):
            v = ws.cell(row, col).value
            if v:
                lines = max(lines, str(v).count("\n") + 1)
        ws.row_dimensions[row].height = min(409, 13 * lines)


def write_many_sheets(pdf_to_pages, out_xlsx: Path) -> Path:
    """
    pdf_to_pages: list of (pdf_path, pages)
      where pages is [(page_number, df), ...]
    """
    wb = Workbook()
    # remove default empty sheet
    if wb.active:
        wb.remove(wb.active)

    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    wrap = Alignment(wrap_text=True, vertical="top")
    used = set()

    for pdf_path, pages in pdf_to_pages:
        logical = []
        for _, df in pages:
            if df.empty:
                continue
            if list(df.columns) == HEADER:
                logical.append(df)
            else:
                logical.append(fold_continuations(df))

        if not logical:
            logical = [pd.DataFrame(columns=HEADER)]

        combined = pd.concat(logical, ignore_index=True)

        ws = wb.create_sheet(title=unique_sheet_name(pdf_path.stem, used))

        # header
        for j, h in enumerate(HEADER, 1):
            cell = ws.cell(1, j, h)
            cell.alignment = wrap
            cell.border = border
        ws.freeze_panes = "A2"

        # data rows
        for i in range(combined.shape[0]):
            for j in range(combined.shape[1]):
                cell = ws.cell(i + 2, j + 1, combined.iat[i, j])
                cell.alignment = wrap
                cell.border = border

        autosize_and_rowheights(ws, EXPECTED_COLS)

    wb.save(out_xlsx)
    return out_xlsx


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--inputs", nargs="*", type=Path, default=[],
                    help="Explicit list of PDF files.")
    ap.add_argument("--dir", type=Path,
                    help="Directory containing PDFs.")
    ap.add_argument("--glob", type=str, default="*.pdf",
                    help="Glob pattern inside --dir (default: *.pdf).")
    ap.add_argument("--out", type=Path, required=True,
                    help="Output .xlsx path.")
    ap.add_argument("--dpi", type=int, default=450,
                    help="Rendering DPI for PDF pages.")
    ap.add_argument("--min-line-frac", type=float, default=0.38,
                    help="Min fraction of page a line must occupy to count as grid.")
    ap.add_argument("--pad", type=int, default=6,
                    help="Padding around detected cells for OCR.")
    ap.add_argument("--psm", type=int, default=6,
                    help="Tesseract PSM mode for OCR.")
    args = ap.parse_args()

    set_tesseract_cmd()

    pdfs = []
    if args.dir:
        pdfs += sorted(args.dir.glob(args.glob))
    if args.inputs:
        pdfs += list(args.inputs)

    # dedupe & keep order
    seen = set()
    ordered = []
    for p in pdfs:
        p = p.resolve()
        if p not in seen:
            seen.add(p)
            ordered.append(p)

    if not ordered:
        sys.exit("No PDFs found.")

    for p in ordered:
        if not p.exists():
            sys.exit(f"Missing: {p}")

    bundle = []
    for p in ordered:
        pages = process_pdf(
            p,
            dpi=args.dpi,
            min_line_frac=args.min_line_frac,
            pad=args.pad,
            psm=args.psm,
        )
        bundle.append((p, pages))
        print(" ->", p.name, "done")

    out = write_many_sheets(bundle, args.out)
    print("Saved workbook:", out)


if __name__ == "__main__":
    main()

"""
Example usage:

python3 testing.py \
  --dir "/Users/helloworld/Downloads/pe2e" \
  --glob "*.pdf" \
  --out "/Users/helloworld/Downloads/pe2e_all.xlsx" \
  --dpi 500 --min-line-frac 0.35 --pad 10 --psm 6
"""
